#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import sys
import re
from pathlib import Path
from typing import Dict, Tuple, List

# Optional deps
try:
    import xlrd  # for .xls read
    from xlutils.copy import copy as xl_copy  # for .xls write
except Exception:
    xlrd = None
    xl_copy = None

try:
    from openpyxl import load_workbook  # for .xlsx read/write
    from openpyxl.utils import column_index_from_string
except Exception:
    load_workbook = None
    column_index_from_string = None


def col_letter_to_idx(letter: str) -> int:
    """Convert column letter (A, B, ..., AA) to 0-based index."""
    letter = letter.strip().upper()
    idx = 0
    for ch in letter:
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f"Invalid column letter: {letter}")
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1


def to_text(v) -> str:
    if v is None:
        return ''
    if isinstance(v, (int,)):
        return str(v)
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    return str(v)


def digits_only(s: str) -> str:
    return ''.join(ch for ch in s if ch.isdigit())


def load_mapping(path: Path, sheet_index: int, id_col_letter: str, name_col_letter: str, skip_header: bool) -> Tuple[Dict[str, str], Dict[str, str]]:
    """Load mapping from id_col -> name_col. Returns (raw_map, digits_map)."""
    ext = path.suffix.lower()
    id_idx = col_letter_to_idx(id_col_letter)
    name_idx = col_letter_to_idx(name_col_letter)

    id_to_name: Dict[str, str] = {}

    if ext == '.xls':
        if xlrd is None:
            raise RuntimeError('xlrd is required for .xls files')
        wb = xlrd.open_workbook(path.as_posix())
        sh = wb.sheet_by_index(sheet_index)
        start_row = 1 if skip_header else 0
        for r in range(start_row, sh.nrows):
            v_id = sh.cell_value(r, id_idx) if id_idx < sh.ncols else ''
            v_name = sh.cell_value(r, name_idx) if name_idx < sh.ncols else ''
            k = to_text(v_id).strip()
            name = to_text(v_name).strip()
            if k:
                id_to_name[k] = name
    elif ext == '.xlsx':
        if load_workbook is None:
            raise RuntimeError('openpyxl is required for .xlsx files')
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[sheet_index]
        # openpyxl is 1-based
        start_row = 2 if skip_header else 1
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            v_id = row[id_idx] if id_idx < len(row) else None
            v_name = row[name_idx] if name_idx < len(row) else None
            k = to_text(v_id).strip()
            name = to_text(v_name).strip()
            if k:
                id_to_name[k] = name
    else:
        raise ValueError('Unsupported mapping file extension: ' + ext)

    id_to_name_digits = {digits_only(k): v for k, v in id_to_name.items() if digits_only(k)}
    return id_to_name, id_to_name_digits


def map_token(token: str, id_to_name: Dict[str, str], id_to_name_digits: Dict[str, str]) -> Tuple[str, bool]:
    # Direct hit
    if token in id_to_name and id_to_name[token] != '':
        return id_to_name[token], True
    # Digits-only
    d = digits_only(token)
    candidates = []
    if d:
        candidates.append(d)
        # Fallback prefixes observed in data (e.g., 6704 -> 66704; 771 -> 66771)
        if len(d) in (3, 4):
            candidates.append('66' + d)
            candidates.append('6' + d)
    for c in candidates:
        if c in id_to_name and id_to_name[c] != '':
            return id_to_name[c], True
        if c in id_to_name_digits and id_to_name_digits[c] != '':
            return id_to_name_digits[c], True
    return token, False


def transform_cell(text: str, keep_prefix: bool, prefix: str, id_to_name: Dict[str, str], id_to_name_digits: Dict[str, str]) -> Tuple[str, int]:
    """Transform cell content like:
    - '物品=ID$1&ID2$2'
    - '220-221$1$80'
    - '7075|80072|523' (uses '|' as segment separator)
    Preserves original segment separators ('&' or '|'). Returns (new_text, unmatched_count).
    """
    if not text:
        return text, 0
    s = to_text(text).strip()
    body = s[len(prefix):] if s.startswith(prefix) else s
    # Split by '&' or '|' while keeping the separators
    tokens = re.split(r'([&|])', body)
    out_parts: List[str] = []
    unmatched = 0

    for tk in tokens:
        if tk in ('&', '|'):
            # Keep original separator
            out_parts.append(tk)
            continue
        if tk == '':
            continue
        seg = tk
        parts = seg.split('$')
        ids_part = parts[0]
        suffix_parts = parts[1:]  # could be none/one/two parts

        id_tokens = [t for t in ids_part.split('-') if t]
        mapped_tokens: List[str] = []
        for t in id_tokens:
            name, hit = map_token(t, id_to_name, id_to_name_digits)
            if not hit:
                unmatched += 1
            mapped_tokens.append(name)
        new_ids = '-'.join(mapped_tokens) if id_tokens else ids_part
        new_seg = new_ids + (('$' + '$'.join(suffix_parts)) if suffix_parts else '')
        out_parts.append(new_seg)

    core = ''.join(out_parts)
    return ((prefix + core) if keep_prefix else core), unmatched


def process_source(source_path: Path, source_sheet_index: int, read_col_letter: str, write_col_letter: str,
                   id_to_name: Dict[str, str], id_to_name_digits: Dict[str, str], keep_prefix: bool, prefix: str,
                   skip_header_source: bool, output_path: Path) -> Tuple[int, int, int]:
    """Process source file, write to output_path. Returns (total_rows, converted_cells, unmatched_count)."""
    ext = source_path.suffix.lower()
    r_idx = col_letter_to_idx(read_col_letter)
    w_idx = col_letter_to_idx(write_col_letter)

    if ext == '.xls':
        if xlrd is None or xl_copy is None:
            raise RuntimeError('xlrd/xlutils are required for .xls files')
        rb = xlrd.open_workbook(source_path.as_posix(), formatting_info=False)
        rs = rb.sheet_by_index(source_sheet_index)
        wb_w = xl_copy(rb)
        ws_w = wb_w.get_sheet(source_sheet_index)

        start_row = 1 if skip_header_source else 0
        converted = 0
        unmatched = 0
        for r in range(start_row, rs.nrows):
            val = rs.cell_value(r, r_idx) if r_idx < rs.ncols else ''
            if not val:
                continue
            new_text, um = transform_cell(val, keep_prefix, prefix, id_to_name, id_to_name_digits)
            ws_w.write(r, w_idx, new_text)
            converted += 1
            unmatched += um
        wb_w.save(output_path.as_posix())
        return rs.nrows, converted, unmatched

    elif ext == '.xlsx':
        if load_workbook is None:
            raise RuntimeError('openpyxl is required for .xlsx files')
        wb = load_workbook(source_path)
        ws = wb.worksheets[source_sheet_index]
        start_row = 2 if skip_header_source else 1
        max_row = ws.max_row
        converted = 0
        unmatched = 0
        for r in range(start_row, max_row + 1):
            cell = ws.cell(row=r, column=w_idx + 1)
            src_cell = ws.cell(row=r, column=r_idx + 1)
            val = src_cell.value
            if val is None:
                continue
            new_text, um = transform_cell(val, keep_prefix, prefix, id_to_name, id_to_name_digits)
            cell.value = new_text
            converted += 1
            unmatched += um
        wb.save(output_path)
        return max_row, converted, unmatched
    else:
        raise ValueError('Unsupported source file extension: ' + ext)


def main():
    ap = argparse.ArgumentParser(description='Map item IDs to names inside concatenated strings, optionally keeping prefix and supporting hyphen-joined IDs.')
    ap.add_argument('--source', required=True, help='Path to source workbook (.xls or .xlsx)')
    ap.add_argument('--source-sheet-index', type=int, default=0, help='Sheet index of source workbook (default: 0)')
    ap.add_argument('--read-col', required=True, help='Column letter to read from, e.g., C')
    ap.add_argument('--write-col', required=True, help='Column letter to write to, e.g., F or H')

    ap.add_argument('--mapping', required=True, help='Path to mapping workbook (.xls or .xlsx)')
    ap.add_argument('--mapping-sheet-index', type=int, default=0, help='Sheet index of mapping workbook (default: 0)')
    ap.add_argument('--id-col', default='A', help='ID column letter in mapping (default: A)')
    ap.add_argument('--name-col', default='B', help='Name column letter in mapping (default: B)')

    ap.add_argument('--keep-prefix', action='store_true', help='Keep the prefix (default: not keep)')
    ap.add_argument('--prefix', default='\u7269\u54c1=', help='Prefix text to preserve/strip (default: 物品=)')

    ap.add_argument('--skip-header-source', action='store_true', help='Skip the first row in source sheet')
    ap.add_argument('--skip-header-mapping', action='store_true', help='Skip the first row in mapping sheet')

    ap.add_argument('--output', help='Explicit output file path. If omitted, a suffix _mapped is added before extension.')

    args = ap.parse_args()

    source_path = Path(args.source)
    mapping_path = Path(args.mapping)

    if not source_path.exists():
        print(f"Source not found: {source_path}", file=sys.stderr)
        sys.exit(1)
    if not mapping_path.exists():
        print(f"Mapping not found: {mapping_path}", file=sys.stderr)
        sys.exit(1)

    id_map, id_map_digits = load_mapping(
        mapping_path,
        sheet_index=args.mapping_sheet_index,
        id_col_letter=args.id_col,
        name_col_letter=args.name_col,
        skip_header=args.skip_header_mapping,
    )

    if args.output:
        out_path = Path(args.output)
    else:
        out_path = source_path.with_name(source_path.stem + '_mapped' + source_path.suffix)

    total_rows, converted_cells, unmatched_count = process_source(
        source_path=source_path,
        source_sheet_index=args.source_sheet_index,
        read_col_letter=args.read_col,
        write_col_letter=args.write_col,
        id_to_name=id_map,
        id_to_name_digits=id_map_digits,
        keep_prefix=args.keep_prefix,
        prefix=args.prefix,
        skip_header_source=args.skip_header_source,
        output_path=out_path,
    )

    print('Source file:', source_path)
    print('Mapping file:', mapping_path)
    print('Output file:', out_path)
    print('Read column:', args.read_col, 'Write column:', args.write_col)
    print('Total rows:', total_rows)
    print('Converted cells:', converted_cells)
    print('Unmatched item IDs count:', unmatched_count)


if __name__ == '__main__':
    main()
