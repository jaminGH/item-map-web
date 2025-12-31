from pathlib import Path
import tempfile
import os
from flask import Flask, render_template, request, send_file, redirect, url_for, flash, send_from_directory, session
from werkzeug.utils import secure_filename
from typing import Dict, Tuple, List
import re

try:
    import xlrd
except Exception:
    xlrd = None

from openpyxl import load_workbook, Workbook

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret-key")

# Persistent storage directories (can be mounted as Docker volumes)
BASE_DATA_DIR = Path(os.environ.get("ITEMMAP_DATA_DIR", "./data")).resolve()
UPLOAD_DIR = BASE_DATA_DIR / "uploads"
OUTPUT_DIR = BASE_DATA_DIR / "outputs"
for d in (UPLOAD_DIR, OUTPUT_DIR):
    d.mkdir(parents=True, exist_ok=True)

# Admin credentials for login form (optional). If set, protect admin routes via session login.
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD")

def require_login(view_func):
    def wrapper(*args, **kwargs):
        if ADMIN_USERNAME and ADMIN_PASSWORD:
            if not session.get('logged_in'):
                flash('请先登录管理员账户')
                return redirect(url_for('login', next=request.path))
        return view_func(*args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


def col_letter_to_idx(letter: str) -> int:
    letter = letter.strip().upper()
    idx = 0
    for ch in letter:
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f"Invalid column letter: {letter}")
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1


def to_text(v) -> str:
    if v is None:
        return ''
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    return str(v)


def digits_only(s: str) -> str:
    return ''.join(ch for ch in s if ch.isdigit())


def load_mapping(path: Path, sheet_index: int, id_col_letter: str, name_col_letter: str, skip_header: bool) -> Tuple[Dict[str, str], Dict[str, str]]:
    ext = path.suffix.lower()
    id_idx = col_letter_to_idx(id_col_letter)
    name_idx = col_letter_to_idx(name_col_letter)
    id_to_name: Dict[str, str] = {}
    if ext == '.xls':
        if xlrd is None:
            raise RuntimeError('xlrd not available to read .xls')
        wb = xlrd.open_workbook(path.as_posix())
        sh = wb.sheet_by_index(sheet_index)
        start_row = 1 if skip_header else 0
        for r in range(start_row, sh.nrows):
            v_id = sh.cell_value(r, id_idx) if id_idx < sh.ncols else ''
            v_name = sh.cell_value(r, name_idx) if name_idx < sh.ncols else ''
            k = to_text(v_id).strip()
            name = to_text(v_name).strip()
            if k:
                id_to_name[k] = name
    elif ext == '.xlsx':
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[sheet_index]
        start_row = 2 if skip_header else 1
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            v_id = row[id_idx] if id_idx < len(row) else None
            v_name = row[name_idx] if name_idx < len(row) else None
            k = to_text(v_id).strip()
            name = to_text(v_name).strip()
            if k:
                id_to_name[k] = name
    else:
        raise ValueError('Unsupported mapping extension')
    id_to_name_digits = {digits_only(k): v for k, v in id_to_name.items() if digits_only(k)}
    return id_to_name, id_to_name_digits


def map_token(token: str, id_to_name: Dict[str, str], id_to_name_digits: Dict[str, str]) -> Tuple[str, bool]:
    # direct hit
    if token in id_to_name and id_to_name[token] != '':
        return id_to_name[token], True
    # digits-only and fallback prefixes
    d = digits_only(token)
    candidates = []
    if d:
        candidates.append(d)
        # fallback prefixes observed in data (e.g., 6704 -> 66704; 771 -> 66771)
        if len(d) in (3, 4):
            candidates.append('66' + d)
            candidates.append('6' + d)
    for c in candidates:
        if c in id_to_name and id_to_name[c] != '':
            return id_to_name[c], True
        if c in id_to_name_digits and id_to_name_digits[c] != '':
            return id_to_name_digits[c], True
    return token, False


def transform_cell(text: str, keep_prefix: bool, prefix: str, id_to_name: Dict[str, str], id_to_name_digits: Dict[str, str], unmatched_acc: List[str] | None = None) -> Tuple[str, int]:
    if not text:
        return text, 0
    s = to_text(text).strip()
    body = s[len(prefix):] if s.startswith(prefix) else s
    # Split by '&' or '|' and keep separators
    tokens = re.split(r'([&|])', body)
    out_parts: List[str] = []
    unmatched = 0
    for tk in tokens:
        if tk in ('&', '|'):
            out_parts.append(tk)
            continue
        if tk == '':
            continue
        seg = tk
        parts = seg.split('$')
        ids_part = parts[0]
        suffix_parts = parts[1:]
        id_tokens = [tok for tok in ids_part.split('-') if tok]
        mapped_tokens: List[str] = []
        for tok in id_tokens:
            name, hit = map_token(tok, id_to_name, id_to_name_digits)
            if not hit:
                unmatched += 1
                if unmatched_acc is not None:
                    unmatched_acc.append(tok)
            mapped_tokens.append(name)
        new_ids = '-'.join(mapped_tokens) if id_tokens else ids_part
        new_seg = new_ids + (('$' + '$'.join(suffix_parts)) if suffix_parts else '')
        out_parts.append(new_seg)
    core = ''.join(out_parts)
    return ((prefix + core) if keep_prefix else core), unmatched


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'GET':
        return render_template('index.html')
    src_file = request.files.get('source')
    map_file = request.files.get('mapping')
    read_col = request.form.get('read_col', '').strip()
    write_col = request.form.get('write_col', '').strip()
    src_sheet_index = int(request.form.get('src_sheet_index', '0') or '0')
    map_sheet_index = int(request.form.get('map_sheet_index', '0') or '0')
    id_col = request.form.get('id_col', 'A').strip() or 'A'
    name_col = request.form.get('name_col', 'B').strip() or 'B'
    keep_prefix = request.form.get('keep_prefix') == 'on'
    prefix = request.form.get('prefix', '物品=').strip() or '物品='
    skip_header_source = request.form.get('skip_header_source') == 'on'
    skip_header_mapping = request.form.get('skip_header_mapping') == 'on'
    export_unmatched = request.form.get('export_unmatched') == 'on'
    if not src_file or not map_file or not read_col or not write_col:
        flash('请提供源文件、映射文件、读取列和回写列')
        return redirect(url_for('index'))
    # Save uploads to persistent dir
    src_name = secure_filename(src_file.filename)
    map_name = secure_filename(map_file.filename)
    src_path = UPLOAD_DIR / src_name
    map_path = UPLOAD_DIR / map_name
    src_file.save(src_path)
    map_file.save(map_path)
    id_to_name, id_to_name_digits = load_mapping(
        map_path,
        map_sheet_index,
        id_col,
        name_col,
        skip_header_mapping,
    )
    ext = src_path.suffix.lower()
    out_path = OUTPUT_DIR / (Path(src_name).stem + '_mapped' + '.xlsx')
    r_idx = col_letter_to_idx(read_col)
    w_idx = col_letter_to_idx(write_col)
    if ext == '.xls':
        if xlrd is None:
            flash('服务器缺少xlrd以读取xls')
            return redirect(url_for('index'))
        rb = xlrd.open_workbook(src_path.as_posix())
        rs = rb.sheet_by_index(src_sheet_index)
        wb_out = Workbook()
        ws_out = wb_out.active
        for r in range(rs.nrows):
            row_vals = []
            for c in range(rs.ncols):
                cell_val = rs.cell_value(r, c)
                row_vals.append(cell_val)
            if r >= (1 if skip_header_source else 0) and r_idx < rs.ncols:
                val = rs.cell_value(r, r_idx)
                if val not in (None, ''):
                    new_text, _ = transform_cell(val, keep_prefix, prefix, id_to_name, id_to_name_digits, unmatched_tokens)
                    while len(row_vals) <= w_idx:
                        row_vals.append('')
                    row_vals[w_idx] = new_text
            ws_out.append(row_vals)
        wb_out.save(out_path.as_posix())
    elif ext == '.xlsx':
        wb = load_workbook(src_path)
        ws = wb.worksheets[src_sheet_index]
        start_row = 2 if skip_header_source else 1
        max_row = ws.max_row
        for r in range(start_row, max_row + 1):
            src_cell = ws.cell(row=r, column=r_idx + 1)
            val = src_cell.value
            if val is None:
                continue
            new_text, _ = transform_cell(val, keep_prefix, prefix, id_to_name, id_to_name_digits, unmatched_tokens)
            ws.cell(row=r, column=w_idx + 1).value = new_text
        wb.save(out_path.as_posix())
    else:
        flash('仅支持 .xls 或 .xlsx 源文件')
        return redirect(url_for('index'))
    # Optionally export unmatched as CSV
    if export_unmatched and unmatched_tokens is not None:
        from collections import Counter
        import csv
        cnt = Counter(unmatched_tokens)
        unmatched_path = OUTPUT_DIR / (Path(src_name).stem + '_unmatched.csv')
        with unmatched_path.open('w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['unmatched_id', 'count'])
            for k, v in sorted(cnt.items(), key=lambda kv: (-kv[1], kv[0])):
                writer.writerow([k, v])
    return send_file(out_path, as_attachment=True, download_name=Path(src_name).stem + '_mapped.xlsx')


@app.route('/admin')
@require_login
def admin():
    uploads = sorted([f.name for f in UPLOAD_DIR.iterdir() if f.is_file()])
    outputs = sorted([f.name for f in OUTPUT_DIR.iterdir() if f.is_file()])
    return render_template('admin.html', uploads=uploads, outputs=outputs)


@app.route('/files/<kind>/<path:filename>')
@require_login
def files(kind: str, filename: str):
    if kind == 'uploads':
        directory = UPLOAD_DIR
    elif kind == 'outputs':
        directory = OUTPUT_DIR
    else:
        flash('未知的文件类别')
        return redirect(url_for('admin'))
    return send_from_directory(directory, filename, as_attachment=True)


@app.route('/settings')
@require_login
def settings():
    info = {
        'BASE_DATA_DIR': str(BASE_DATA_DIR),
        'UPLOAD_DIR': str(UPLOAD_DIR),
        'OUTPUT_DIR': str(OUTPUT_DIR),
        'PORT': os.environ.get('PORT', '8000'),
        'ADMIN_ENABLED': bool(ADMIN_USERNAME and ADMIN_PASSWORD),
    }
    return render_template('settings.html', info=info)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if not (ADMIN_USERNAME and ADMIN_PASSWORD):
        flash('未配置管理员账户，当前无需登录')
        return redirect(url_for('admin'))
    next_url = request.args.get('next') or url_for('admin')
    if request.method == 'POST':
        user = request.form.get('username', '')
        pw = request.form.get('password', '')
        if user == ADMIN_USERNAME and pw == ADMIN_PASSWORD:
            session['logged_in'] = True
            flash('登录成功')
            return redirect(next_url)
        flash('用户名或密码错误')
    return render_template('login.html', next_url=next_url)


@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    flash('已退出登录')
    return redirect(url_for('index'))


if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5000, debug=True)
